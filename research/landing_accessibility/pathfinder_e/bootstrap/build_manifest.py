"""E Pathfinder P0 bootstrap — SSOTV3 parsing + route-work manifest + hash inventory.

절대 하지 않는 것 (P0/OFFLINE 경계):
- mobile_web_eligibility 를 직접 판정하지 않는다 (URL 을 실제로 열어야 하는 precheck 활동이라
  REAL 이며, P0 는 OFFLINE 전용이다). 원본 값(PRECHECK_REQUIRED)을 그대로 보존한다.
- task family/endpoint/target 을 바꾸지 않는다. SSOTV3 원문을 구조화할 뿐이다.
- 여기서 계산하는 해시는 A 의 공식 freeze 해시가 아니다 — E 가 현재 candidate 데이터 위에서
  계산한 작업용 참조 해시(`e_working_*`)다. A 가 실제 freeze 를 발행하면 그 해시로 대조한다.

입력: /home/sieg/projects-wsl/ProjectFinal/SSOTV3/*  (절대경로 고정 — 실행 위치와 무관)
출력: 이 스크립트가 있는 디렉터리(pathfinder_e/bootstrap/)의 형제 파일들.
"""

from __future__ import annotations

import csv
import hashlib
import json
import unicodedata
from pathlib import Path

import openpyxl

SSOT = Path("/home/sieg/projects-wsl/ProjectFinal/SSOTV3")
OUT = Path(__file__).resolve().parent

CSV_PATH = SSOT / "CROSS_SERVICE_TASK_REGISTRY_50_v3.0.csv"
XLSX_PATH = SSOT / "CROSS_SERVICE_FLOW_PIVOT_v3.0.xlsx"
CANDIDATE_JSON_PATH = SSOT / "CROSS_SERVICE_TARGET_FRAME_50_v3.0_candidate.json"
MANIFEST_META_PATH = SSOT / "MANIFEST_v3.0.json"

# ── 안전 규약 (내 역할 §5) — SSOTV3 원문에 itemized list 로 존재하지 않아 직접 손으로 인코딩한다.
# NLP/자동추출로 만들지 않는다: 이 목록 자체가 안전경계이므로 원문 grep 이 아니라 직접 읽고 옮겨적은
# 값이어야 한다 (SSOT 가 반복 경고하는 "화면/텍스트에서 자동추론" 함정을 안전목록 자체에 재현하지 않기 위함).
GLOBAL_FORBIDDEN = [
    "credential_input",
    "login_submit",
    "otp_or_identity_verification",
    "captcha_solve_or_bypass",
    "real_money_transfer",
    "cart_add",
    "purchase_or_order_or_reservation",
    "seat_selection",
    "payment",
    "phone_call_connect",
    "external_app_launch",
    "location_permission_grant",
    "real_personal_or_tracking_or_account_or_user_info_input",
    "terms_agreement_or_signup_completion",
    "account_creation_to_view_results",
]

# family_id -> family-specific forbidden actions, 01_TASK_FAMILY_TARGET_FRAME_v3.0.md §2 / 00 §4 원문 그대로 옮김
FAMILY_FORBIDDEN = {
    "F1": [
        "credential_input",
        "login_submit",
        "recipient_account_input",
        "amount_input",
        "transfer_execution",
    ],
    "F2": ["cart_activation", "purchase_activation", "checkout_activation", "payment_activation"],
    "F3": [
        "real_tracking_number_input",
        "personal_data_input",
        "query_submit",  # endpoint 자체가 "조회 실행 control 이 관측 가능한 최초 상태" — submit 이전에 멈춘다
    ],
    "F4": ["reservation", "phone_call", "external_app_launch", "location_permission_grant"],
    "F5": ["seat_selection", "reservation", "payment"],
}

# 04_FLOW_CODEBOOK_v3.0.md 의 17 canonical action token (ABSTAIN 포함) — XLSX 04_FLOW_CODEBOOK 시트와 대조검증
EXPECTED_TOKENS = [
    "OPEN_GLOBAL_MENU", "OPEN_LOCAL_MENU", "SWITCH_TAB", "EXPAND_ACCORDION",
    "SELECT_CATEGORY", "SELECT_FUNCTION", "INPUT_QUERY", "SELECT_ORIGIN",
    "SELECT_DESTINATION", "SELECT_DATE", "SUBMIT_QUERY", "SELECT_RESULT",
    "OPEN_ITEM_DETAIL", "OPEN_PLACE_DETAIL", "DISMISS_OBSTRUCTION", "AUTH_GATE",
    "ENDPOINT_REACHED", "ABSTAIN",
]


def norm(text: str | None) -> str:
    if text is None:
        return ""
    return unicodedata.normalize("NFC", " ".join(text.split()))


def sha256_text(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def sha256_json(obj) -> str:
    payload = json.dumps(obj, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return sha256_text(payload)


def load_csv_targets() -> list[dict]:
    with CSV_PATH.open(encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def load_xlsx_families() -> list[dict]:
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb["01_TASK_FAMILIES"]
    rows = list(ws.iter_rows(values_only=True))
    header, body = rows[0], rows[1:]
    return [dict(zip(header, row)) for row in body if row[0]]


def load_xlsx_tokens() -> list[str]:
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb["04_FLOW_CODEBOOK"]
    rows = list(ws.iter_rows(values_only=True))
    return [r[0] for r in rows[1:] if r[0]]


def load_candidate_json() -> dict:
    return json.loads(CANDIDATE_JSON_PATH.read_text(encoding="utf-8"))


def main() -> None:
    csv_targets = load_csv_targets()
    families = load_xlsx_families()
    tokens = load_xlsx_tokens()
    candidate = load_candidate_json()
    manifest_meta = json.loads(MANIFEST_META_PATH.read_text(encoding="utf-8"))

    qa = {"checks": [], "status": "PASS"}

    def check(name: str, ok: bool, detail: str = "") -> None:
        qa["checks"].append({"check": name, "ok": ok, "detail": detail})
        if not ok:
            qa["status"] = "FAIL"

    # ── QA 1: CSV / XLSX / candidate JSON target_id 집합 일치
    csv_ids = {r["target_id"] for r in csv_targets}
    json_ids = {t["target_id"] for t in candidate["targets"]}
    check("csv_row_count_50", len(csv_targets) == 50, f"csv rows={len(csv_targets)}")
    check("csv_json_target_id_match", csv_ids == json_ids,
          f"csv-only={sorted(csv_ids - json_ids)} json-only={sorted(json_ids - csv_ids)}")
    check("family_count_5", len(families) == 5, f"families={[f['family_id'] for f in families]}")
    check("token_count_18", len(tokens) == 18, f"xlsx tokens={tokens}")
    check("token_set_matches_role_spec", set(tokens) == set(EXPECTED_TOKENS),
          f"diff={set(tokens) ^ set(EXPECTED_TOKENS)}")
    for fam in families:
        n_csv = sum(1 for r in csv_targets if r["family_id"] == fam["family_id"])
        check(f"family_{fam['family_id']}_n10", n_csv == 10, f"csv n={n_csv}")

    # ── family-level 해시 (endpoint_contract, task contract 기본형)
    family_by_id = {f["family_id"]: f for f in families}
    family_hashes = {}
    for fid, fam in family_by_id.items():
        endpoint_norm = norm(fam["endpoint_contract"])
        endpoint_hash = sha256_text(endpoint_norm)
        contract_obj = {
            "family_id": fid,
            "task_family": fam["task_family"],
            "matched_task": norm(fam["matched_task"]),
            "task_instruction": norm(fam["task_instruction"]),
            "fixed_fixture": norm(fam["fixed_fixture"]),
            "endpoint_contract": endpoint_norm,
            "global_forbidden": GLOBAL_FORBIDDEN,
            "family_forbidden": FAMILY_FORBIDDEN[fid],
        }
        family_hashes[fid] = {
            "e_working_endpoint_contract_hash": endpoint_hash,
            "e_working_task_contract_hash": sha256_json(contract_obj),
            "hash_input": contract_obj,
        }

    # ── target-level manifest row 구성 (fixture_override 있으면 target-specific task_contract_hash)
    route_work_targets = []
    for row in csv_targets:
        fid = row["family_id"]
        fam_hash = family_hashes[fid]
        fixture_override = norm(row.get("fixture_override") or "") or None
        if fixture_override:
            contract_obj = dict(fam_hash["hash_input"])
            contract_obj["fixture_override"] = fixture_override
            contract_obj["target_id"] = row["target_id"]
            task_contract_hash = sha256_json(contract_obj)
        else:
            task_contract_hash = fam_hash["e_working_task_contract_hash"]

        route_work_targets.append({
            "target_id": row["target_id"],
            "family_id": fid,
            "task_family": row["task_family"],
            "legacy_archetype": row["legacy_archetype"],
            "service_name": row["service_name"],
            "provider_type": row["provider_type"],
            "official_entry_url": row["official_entry_url"],
            "matched_task": row["matched_task"],
            "task_instruction": row["task_instruction"],
            "fixed_fixture": row["fixed_fixture"] or None,
            "fixture_override": fixture_override,
            "endpoint_contract": row["endpoint_contract"],
            "forbidden_actions": {
                "global": GLOBAL_FORBIDDEN,
                "family_specific": FAMILY_FORBIDDEN[fid],
            },
            "e_working_task_contract_hash": task_contract_hash,
            "e_working_endpoint_contract_hash": fam_hash["e_working_endpoint_contract_hash"],
            "mobile_web_eligibility": row["mobile_web_eligibility"],
            "mobile_web_eligibility_note": (
                "SSOTV3 원본값 그대로 보존. E 가 판정하지 않음 — 실제 판정은 "
                "URL 을 여는 REAL precheck 이며 P0/OFFLINE 범위 밖."
            ),
            "task_freeze_status": row["task_freeze_status"],
            "scout_status": "NOT_STARTED",
            "scout_status_note": "A 가 REAL E_SCOUT scope 를 열기 전까지 유지되는 상태값.",
        })

    # target_id 순서 고정 — outcome 기대치로 재정렬 금지 (내 역할 §11)
    route_work_targets.sort(key=lambda t: t["target_id"])

    manifest = {
        "manifest_kind": "E_ROUTE_WORK_MANIFEST",
        "authority_status": "AUXILIARY_EXECUTION_EVIDENCE",
        "canonical": False,
        "self_approved": False,
        "generated_by": "claude-e/pathfinder-v3",
        "ssot_source": "SSOTV3 (AUTHORITY_CANDIDATE / NO_NEW_REAL_TARGET_RELEASE)",
        "ssot_pack_id": manifest_meta["pack_id"],
        "ssot_pack_status": manifest_meta["status"],
        "candidate_frame_id": candidate["frame_id"],
        "candidate_real_target_allowed": candidate["real_target_allowed"],
        "note": (
            "이 manifest 는 E 가 SSOTV3 candidate 데이터를 구조화한 작업용 산출물이다. "
            "target 순서/구성을 바꾸는 권한은 없다. real_target_allowed=false — "
            "A 가 별도 REAL E_SCOUT scope 를 열기 전까지 어떤 target 도 실제로 열지 않는다."
        ),
        "ordering_rule": "target_id lexicographic — outcome 기대치로 재정렬 금지",
        "target_count": len(route_work_targets),
        "family_count": len(families),
        "targets": route_work_targets,
    }

    task_contract_inventory = {
        "kind": "E_TASK_CONTRACT_INVENTORY",
        "authority_status": "AUXILIARY_EXECUTION_EVIDENCE",
        "canonical": False,
        "note": "family-level 해시 계보 + target-level override 계보. A freeze 전 참조용.",
        "hash_scheme": {
            "endpoint_contract_hash": "sha256(NFC-normalize + whitespace-collapse(endpoint_contract text))",
            "task_contract_hash": (
                "sha256(canonical-json({family_id, task_family, matched_task, task_instruction, "
                "fixed_fixture, endpoint_contract, global_forbidden, family_forbidden[, fixture_override, "
                "target_id if override present]}, sort_keys=True, ensure_ascii=False))"
            ),
        },
        "global_forbidden_actions": GLOBAL_FORBIDDEN,
        "families": {
            fid: {
                "task_family": family_by_id[fid]["task_family"],
                "domain": family_by_id[fid]["domain"],
                "legacy_archetype": family_by_id[fid]["legacy_archetype"],
                "n": family_by_id[fid]["n"],
                "family_specific_forbidden_actions": FAMILY_FORBIDDEN[fid],
                **{k: v for k, v in family_hashes[fid].items() if k != "hash_input"},
            }
            for fid in family_by_id
        },
        "targets_with_fixture_override": [
            t["target_id"] for t in route_work_targets if t["fixture_override"]
        ],
    }

    qa["token_inventory"] = {"xlsx_tokens": tokens, "role_spec_tokens": EXPECTED_TOKENS}

    (OUT / "ROUTE_WORK_MANIFEST.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (OUT / "TASK_CONTRACT_INVENTORY.json").write_text(
        json.dumps(task_contract_inventory, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (OUT / "PARSE_QA_REPORT.json").write_text(
        json.dumps(qa, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    print(f"QA status: {qa['status']}")
    for c in qa["checks"]:
        mark = "OK" if c["ok"] else "FAIL"
        print(f"  [{mark}] {c['check']} {c['detail']}")
    print(f"\ntargets_with_fixture_override: {task_contract_inventory['targets_with_fixture_override']}")
    print(f"wrote: {OUT / 'ROUTE_WORK_MANIFEST.json'}")
    print(f"wrote: {OUT / 'TASK_CONTRACT_INVENTORY.json'}")
    print(f"wrote: {OUT / 'PARSE_QA_REPORT.json'}")


if __name__ == "__main__":
    main()
